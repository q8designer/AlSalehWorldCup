#!/usr/bin/env python3
"""
World Cup 2026 Score Fetcher & Excel Updater
"""

import os, sys, argparse, requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── Always run from the script's own directory ─────────────────────────────────
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ─── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH    = "WorldCup2026_Competition.xlsx"
MATCHES_SHEET = "📋 Matches"

API_KEY = os.environ.get("FOOTBALL_API_KEY", "YOUR_API_KEY_HERE")
API_BASE         = "https://api.football-data.org/v4"
COMPETITION_CODE = "WC"

# ─── Team name aliases ─────────────────────────────────────────────────────────
TEAM_ALIASES = {
    "USA":                        "United States",
    "United States of America":   "United States",
    "Türkiye":                    "Turkey",
    "Korea Republic":             "South Korea",
    "IR Iran":                    "Iran",
    "Bosnia-Herzegovina":         "Bosnia and Herzeg",
    "Czechia":                    "Czech Republic",
    "DR Congo":                   "Congo DR",
    "Curaçao":                    "Curacao",
}

def norm(name):
    return TEAM_ALIASES.get(name, name).strip().lower()

def norm_match(home, away):
    return f"{norm(home)} vs {norm(away)}"

# ─── Manual scores ─────────────────────────────────────────────────────────────
def manual_scores():
    return [
        ("Mexico",        "South Africa",    2, 0),
        ("South Korea",   "Czech Republic",  2, 1),
        ("Canada",        "Bosnia and Herzeg", 1, 1),
        ("United States", "Paraguay",        4, 1),
        # Add more here as matches finish:
    ]

# ─── API fetch ─────────────────────────────────────────────────────────────────
def fetch_wc_scores():
    headers = {"X-Auth-Token": API_KEY} if API_KEY != "YOUR_API_KEY_HERE" else {}
    try:
        url  = f"{API_BASE}/competitions/{COMPETITION_CODE}/matches?status=FINISHED"
        resp = requests.get(url, headers=headers, timeout=15)

        if resp.status_code == 401:
            print("⚠️  API key missing or invalid.")
            print("    Get a free key: https://www.football-data.org/client/register")
            return []
        if resp.status_code == 404:
            print("⚠️  World Cup 2026 data not yet available. Try after June 11.")
            return []

        resp.raise_for_status()
        results = []
        for m in resp.json().get("matches", []):
            s = m["score"]["fullTime"]
            hg, ag = s.get("home"), s.get("away")
            if hg is None or ag is None:
                continue
            results.append({
                "home": m["homeTeam"]["name"],
                "away": m["awayTeam"]["name"],
                "home_goals": int(hg),
                "away_goals": int(ag),
            })
        print(f"✅ Fetched {len(results)} completed matches from API")
        return results

    except requests.exceptions.ConnectionError:
        print("❌ No internet connection.")
        return []
    except Exception as e:
        print(f"❌ API error: {e}")
        return []

# ─── Excel updater ─────────────────────────────────────────────────────────────
def load_excel_matches(wb):
    ws = wb[MATCHES_SHEET]
    matches = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row[1].value:
            continue
        home_team = row[3].value
        away_team = row[7].value
        key = norm_match(home_team or "", away_team or "")
        matches[key] = {
            "row":          row[1].row,
            "home_cell":    row[4],
            "away_cell":    row[5],
            "status_cell":  row[8],
            "current_home": row[4].value,
            "current_away": row[5].value,
        }
    return matches

def update_excel(results, dry_run=False):
    wb      = load_workbook(EXCEL_PATH)
    ex      = load_excel_matches(wb)
    updated = skipped = not_found = 0

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Updating scores...\n")
    print(f"{'Match':<48} {'Score':^9} Status")
    print("─" * 72)

    for r in results:
        key = norm_match(r["home"], r["away"])
        if key not in ex:
            key_rev = norm_match(r["away"], r["home"])
            if key_rev not in ex:
                print(f"  ⚠️  Not matched: {r['home']} vs {r['away']}")
                not_found += 1
                continue
            key = key_rev
            r["home_goals"], r["away_goals"] = r["away_goals"], r["home_goals"]

        m     = ex[key]
        score = f"{r['home_goals']} - {r['away_goals']}"
        label = f"{r['home']} vs {r['away']}"

        if m["current_home"] == r["home_goals"] and m["current_away"] == r["away_goals"]:
            print(f"  ✓  {label:<44} {score:^9} Already up-to-date")
            skipped += 1
        else:
            if not dry_run:
                m["home_cell"].value   = r["home_goals"]
                m["away_cell"].value   = r["away_goals"]
                m["status_cell"].value = "✅ Done"
                m["status_cell"].fill  = PatternFill("solid", fgColor="D5F5E3")
            prev = f"(was {m['current_home']}-{m['current_away']})" if m["current_home"] is not None else "(new)"
            print(f"  📝 {label:<44} {score:^9} Updated {prev}")
            updated += 1

    print("─" * 72)
    print(f"\n  Updated: {updated}  |  Current: {skipped}  |  Unmatched: {not_found}")

    if not dry_run and updated > 0:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        wb[MATCHES_SHEET].cell(row=1, column=10, value=f"Last updated: {ts}")
        wb.save(EXCEL_PATH)
        print(f"\n✅ Saved → {EXCEL_PATH}")
    elif dry_run:
        print("\n  (Dry run — nothing written)")
    else:
        print("\n  No updates needed.")

def list_matches():
    wb = load_workbook(EXCEL_PATH)
    ws = wb[MATCHES_SHEET]
    print(f"\n{'Rnd':<5} {'Match':<48} {'Score':^9} Status")
    print("─" * 78)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        score = f"{row[4]} - {row[5]}" if row[4] is not None else "  -  "
        print(f"{str(row[0]):<5} {str(row[1]):<48} {score:^9} {row[8] or '⏳ Pending'}")

# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="World Cup 2026 Score Updater")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--list",    action="store_true")
    parser.add_argument("--manual",  action="store_true")
    args = parser.parse_args()

    print("⚽ World Cup 2026 Score Updater")
    print(f"   Working directory: {os.getcwd()}")
    print(f"   Excel file: {EXCEL_PATH}")
    print(f"   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    if args.list:
        list_matches()
        return

    if args.manual:
        raw     = manual_scores()
        results = [{"home":h,"away":a,"home_goals":hg,"away_goals":ag} for h,a,hg,ag in raw]
        print(f"📋 Using {len(results)} manual scores")
    else:
        results = fetch_wc_scores()
        if not results:
            print("\n💡 Use --manual to enter scores directly in this script.")
            return

    update_excel(results, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
