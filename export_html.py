#!/usr/bin/env python3
"""
World Cup 2026 – HTML Leaderboard Exporter
"""

import json, argparse, base64, os
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ── Always run from the script's own directory ─────────────────────────────────
os.chdir(os.path.dirname(os.path.abspath(__file__)))

EXCEL_PATH = "WorldCup2026_Competition.xlsx"

def medal(rank):
    return {1:"🥇", 2:"🥈", 3:"🥉"}.get(rank, str(rank))

def generate_html(excel_path, out_path):
    # Load logo
    logo_path = os.path.join(os.path.dirname(os.path.abspath(excel_path)), "IMG_8424.jpg")
    logo_b64 = ""
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as lf:
            logo_b64 = base64.b64encode(lf.read()).decode()
    logo_tag = f"data:image/jpeg;base64,{logo_b64}" if logo_b64 else ""

    wb = load_workbook(excel_path, data_only=True)

    ws_m = wb["📋 Matches"]
    matches = []
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        matches.append({
            "round": row[0], "match": row[1],
            "home_team": row[3], "home_score": row[4],
            "away_score": row[5], "away_team": row[7], "status": row[8],
        })

    ws_p = wb["📊 Predictions"]
    totals = defaultdict(lambda: {"total":0,"win":0,"home":0,"away":0,"diff":0,"scored":0})
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        n = row[1]
        totals[n]["total"] += row[11] if isinstance(row[11],(int,float)) else 0
        totals[n]["win"]   += row[7]  if isinstance(row[7], (int,float)) else 0
        totals[n]["home"]  += row[8]  if isinstance(row[8], (int,float)) else 0
        totals[n]["away"]  += row[9]  if isinstance(row[9], (int,float)) else 0
        totals[n]["diff"]  += row[10] if isinstance(row[10],(int,float)) else 0
        if row[5] is not None:
            totals[n]["scored"] += 1

    leaderboard = sorted(totals.items(), key=lambda x: x[1]["total"], reverse=True)
    completed   = [m for m in matches if m["home_score"] is not None]
    pending     = [m for m in matches if m["home_score"] is None]
    timestamp   = datetime.now().strftime("%d %b %Y, %H:%M")

    lb_rows = ""
    for i, (name, d) in enumerate(leaderboard, 1):
        cls = "gold" if i==1 else "silver" if i==2 else "bronze" if i==3 else ("even" if i%2==0 else "odd")
        lb_rows += f"""
        <tr class="{cls}">
          <td class="rank">{medal(i)}</td>
          <td class="name">{name}</td>
          <td class="pts total">{int(d['total'])}</td>
          <td class="pts">{int(d['win'])}</td>
          <td class="pts">{int(d['home'])}</td>
          <td class="pts">{int(d['away'])}</td>
          <td class="pts">{int(d['diff'])}</td>
          <td class="pts muted">{int(d['scored'])}</td>
        </tr>"""

    match_rows = ""
    for m in matches:
        done = m["home_score"] is not None
        score = f"{m['home_score']} – {m['away_score']}" if done else "vs"
        match_rows += f"""
        <tr class="{'done' if done else 'pending'}">
          <td>{m['round'] or ''}</td>
          <td>{m['home_team'] or ''}</td>
          <td class="score-cell">{score}</td>
          <td>{m['away_team'] or ''}</td>
          <td><span class="{'badge-done' if done else 'badge-pending'}">{'✅ Done' if done else '⏳ Pending'}</span></td>
        </tr>"""

    logo_img = f'<img class="hero-logo" src="{logo_tag}" alt="Alsaleh World Cup Logo">' if logo_tag else ""

    html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Alsaleh World Cup Tournament 2026</title>
<style>
:root{{--bg:#0d1117;--surface:#161b22;--border:#30363d;--accent:#c0392b;--gold:#f5a623;--text:#e6edf3;--muted:#8b949e;--green:#3fb950;}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;}}
.hero{{background:linear-gradient(135deg,#1a0000,#2d0000,#1a0a00);padding:0;border-bottom:3px solid #c0392b;}}
.hero-inner{{display:flex;align-items:center;justify-content:center;gap:32px;padding:32px 24px 28px;flex-wrap:wrap;}}
.hero-logo{{width:130px;height:130px;object-fit:contain;filter:drop-shadow(0 4px 16px rgba(192,57,43,0.5));flex-shrink:0;}}
.hero-eyebrow{{font-size:.75rem;letter-spacing:3px;text-transform:uppercase;color:var(--gold);font-weight:600;margin-bottom:6px;}}
.hero-title{{font-size:clamp(2rem,5vw,3.2rem);font-weight:900;line-height:1.05;letter-spacing:-1px;}}
.hero-title .alsaleh{{background:linear-gradient(135deg,#8B0000,#c0392b,#e74c3c);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;display:block;}}
.hero-title .wct{{color:#fff;display:block;font-size:0.62em;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-top:2px;}}
.hero-sub{{color:var(--muted);margin-top:10px;font-size:.9rem;}}
.hero-badge{{display:inline-block;background:rgba(192,57,43,.2);border:1px solid rgba(192,57,43,.4);border-radius:20px;padding:3px 12px;font-size:.75rem;color:#e74c3c;font-weight:600;margin-top:8px;letter-spacing:1px;}}
.stats-bar{{display:flex;justify-content:center;gap:40px;flex-wrap:wrap;padding:18px 24px;background:var(--surface);border-bottom:1px solid var(--border);}}
.stat{{text-align:center;}}
.stat .num{{font-size:1.8rem;font-weight:700;color:var(--gold);}}
.stat .lbl{{font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:1.5px;margin-top:2px;}}
.tabs{{display:flex;background:var(--surface);border-bottom:1px solid var(--border);padding:0 24px;}}
.tab{{padding:14px 24px;cursor:pointer;border-bottom:3px solid transparent;font-weight:600;font-size:.9rem;color:var(--muted);transition:all .2s;user-select:none;}}
.tab.active{{color:var(--accent);border-color:var(--accent);}}
.tab:hover{{color:var(--text);}}
.content{{max-width:1100px;margin:0 auto;padding:24px 16px;}}
.panel{{display:none;}}.panel.active{{display:block;}}
.table-wrap{{overflow-x:auto;border-radius:12px;border:1px solid var(--border);}}
table{{width:100%;border-collapse:collapse;font-size:.9rem;}}
thead th{{background:#1a0000;color:var(--muted);font-size:.72rem;text-transform:uppercase;letter-spacing:1px;padding:13px 16px;text-align:center;white-space:nowrap;}}
thead th:nth-child(2){{text-align:left;}}
tbody tr{{border-bottom:1px solid var(--border);transition:background .15s;}}
tbody tr:hover{{background:rgba(192,57,43,.06);}}
td{{padding:12px 16px;text-align:center;color:#111;}}
td.name{{text-align:left;font-weight:600;color:#000;}}
td.rank{{font-size:1.1rem;font-weight:700;width:52px;}}
td.total{{font-weight:800;font-size:1.05rem;color:var(--accent);}}
td.muted{{color:#555;}}
tr.gold{{background:rgba(245,166,35,.12);}}
tr.silver{{background:rgba(192,192,192,.08);}}
tr.bronze{{background:rgba(205,127,50,.08);}}
tr.even{{background:rgba(255,255,255,.02);}}
tr.pending{{opacity:.6;}}
td.score-cell{{font-size:1rem;font-weight:800;color:var(--accent);width:90px;}}
.badge-done{{display:inline-block;background:rgba(63,185,80,.15);color:var(--green);border-radius:6px;padding:2px 10px;font-size:.8rem;}}
.badge-pending{{display:inline-block;background:rgba(255,255,255,.06);color:var(--muted);border-radius:6px;padding:2px 10px;font-size:.8rem;}}
.points-key{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px;padding:14px 16px;background:var(--surface);border-radius:10px;border:1px solid var(--border);align-items:center;}}
.pk{{display:flex;align-items:center;gap:8px;font-size:.85rem;}}
.pk-badge{{border-radius:6px;padding:2px 9px;font-weight:700;font-size:.8rem;color:#fff;}}
.b48{{background:#8B0000;}}.b15{{background:#2980b9;}}.b12{{background:#27ae60;}}
.search-bar{{margin-bottom:16px;}}
.search-bar input{{width:100%;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 14px;color:var(--text);font-size:.9rem;outline:none;transition:border-color .2s;}}
.search-bar input:focus{{border-color:var(--accent);}}
footer{{text-align:center;padding:32px;color:var(--muted);font-size:.8rem;border-top:1px solid var(--border);margin-top:16px;}}
@media(max-width:600px){{td{{padding:10px 8px;font-size:.8rem;color:#111;}}}}
</style></head><body>
<div class="hero">
  <div class="hero-inner">
    {logo_img}
    <div class="hero-text">
      <div class="hero-eyebrow">⚽ Official Leaderboard</div>
      <h1 class="hero-title">
        <span class="alsaleh">ALSALEH</span>
        <span class="wct">World Cup Tournament</span>
      </h1>
      <div class="hero-sub">Group stage predictions — updated after each match</div>
      <div class="hero-badge">FIFA World Cup 2026</div>
    </div>
  </div>
</div>
<div class="stats-bar">
  <div class="stat"><div class="num">{len(leaderboard)}</div><div class="lbl">Participants</div></div>
  <div class="stat"><div class="num">{len(matches)}</div><div class="lbl">Total Matches</div></div>
  <div class="stat"><div class="num">{len(completed)}</div><div class="lbl">Completed</div></div>
  <div class="stat"><div class="num">{len(pending)}</div><div class="lbl">Remaining</div></div>
  <div class="stat"><div class="num">{leaderboard[0][1]['total']}</div><div class="lbl">Top Score</div></div>
</div>
<div class="tabs">
  <div class="tab active" onclick="switchTab('leaderboard',this)">🏆 Leaderboard</div>
  <div class="tab" onclick="switchTab('matches',this)">📋 Matches</div>
</div>
<div class="content">
  <div id="leaderboard" class="panel active">
    <div class="points-key">
      <span style="font-size:.8rem;color:var(--muted)">Points:</span>
      <div class="pk"><span class="pk-badge b48">48</span>Correct winner</div>
      <div class="pk"><span class="pk-badge b15">15</span>Exact home goals</div>
      <div class="pk"><span class="pk-badge b15">15</span>Exact away goals</div>
      <div class="pk"><span class="pk-badge b12">12</span>Exact goal difference</div>
    </div>
    <div class="search-bar"><input type="text" placeholder="🔍 Search participant..." oninput="filterTable(this.value)"></div>
    <div class="table-wrap">
      <table><thead><tr>
        <th>#</th><th style="text-align:left">Participant</th><th>Total Pts</th>
        <th>Win (48)</th><th>Home (15)</th><th>Away (15)</th><th>Diff (12)</th><th>Matches</th>
      </tr></thead>
      <tbody id="lbBody">{lb_rows}</tbody></table>
    </div>
  </div>
  <div id="matches" class="panel">
    <div class="table-wrap">
      <table><thead><tr><th>Round</th><th>Home</th><th>Score</th><th>Away</th><th>Status</th></tr></thead>
      <tbody>{match_rows}</tbody></table>
    </div>
  </div>
</div>
<footer>Last updated: {timestamp} &nbsp;·&nbsp; Alsaleh World Cup Tournament 2026</footer>
<script>
function switchTab(id,el){{
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
}}
function filterTable(q){{
  q=q.toLowerCase();
  document.querySelectorAll('#lbBody tr').forEach(r=>{{
    r.style.display=r.querySelector('.name')?.textContent.toLowerCase().includes(q)?'':'none';
  }});
}}
</script>
</body></html>"""

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅ HTML saved → {out_path}")
    print(f"   {len(leaderboard)} participants | {len(completed)}/{len(matches)} matches completed")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="leaderboard.html")
    args = parser.parse_args()
    generate_html(EXCEL_PATH, args.out)
